"""Grade a multi-system comparison XLSX.

Reads an Excel workbook where each sheet (except 'readme') contains
columns: question, expectedResponse, actualResponse.  Grades every row
with the same LLM-judge rubric used by grade_qa_answers.py and writes
an output XLSX with grading columns appended to each sheet.

Usage:
    python yar/evaluation/grade_comparison.py --input mindhub_eval.xlsx
    python yar/evaluation/grade_comparison.py --input eval.xlsx --judge-model openai/gpt-4.1-mini
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

from yar.evaluation.grade_qa_answers import (
    grade_row,
    normalize_litellm_model_name,
)
from yar.evaluation.qa_eval_common import (
    DEFAULT_JUDGE_API_BASE,
    DEFAULT_JUDGE_API_KEY,
    DEFAULT_YAR_API_KEY,
    DEFAULT_YAR_API_URL,
    RESULTS_DIR,
    fetch_yar_health,
    timestamped_results_path,
)

SKIP_SHEETS = {'readme'}
REQUIRED_COLUMNS = {'question', 'actualResponse'}
GRADE_COLUMNS = [
    'generalQuality',
    'seemsRelevant',
    'seemsComplete',
    'basedOnKnowledgeSources',
    'reason',
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Grade a multi-system comparison XLSX with an LLM judge.',
    )
    parser.add_argument('--input', '-i', type=Path, required=True, help='Input XLSX file')
    parser.add_argument('--output', '-o', type=Path, default=None, help='Output XLSX (default: results/comparison_grades_<ts>.xlsx)')
    parser.add_argument('--judge-api-base', type=str, default=DEFAULT_JUDGE_API_BASE)
    parser.add_argument('--judge-api-key', type=str, default=DEFAULT_JUDGE_API_KEY)
    parser.add_argument('--judge-model', type=str, default='')
    parser.add_argument('--yar-api-url', type=str, default=DEFAULT_YAR_API_URL)
    parser.add_argument('--yar-api-key', type=str, default=DEFAULT_YAR_API_KEY)
    parser.add_argument('--temperature', type=float, default=0.0)
    parser.add_argument('--max-tokens', type=int, default=300)
    parser.add_argument('--allow-inline-citations', action='store_true')
    parser.add_argument('--sheets', type=str, default=None, help='Comma-separated sheet names to grade (default: all non-readme)')
    return parser.parse_args()


def read_sheet_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, str]]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {h: (str(v) if v is not None else '') for h, v in zip(headers, row, strict=False)}
        if not record.get('question', '').strip():
            continue
        rows.append(record)
    return rows


def write_graded_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    rows: list[dict[str, str]],
) -> None:
    """Create a new sheet in wb with original data + grade columns."""
    ws = wb.create_sheet(title=sheet_name)
    if not rows:
        return

    all_keys = list(rows[0].keys())
    for col_idx, key in enumerate(all_keys, start=1):
        ws.cell(row=1, column=col_idx, value=key)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))


def grade_sheet(
    sheet_name: str,
    rows: list[dict[str, str]],
    judge_model: str,
    judge_api_base: str,
    judge_api_key: str,
    temperature: float,
    max_tokens: int,
    allow_inline_citations: bool,
) -> list[dict[str, str]]:
    """Grade every row in a sheet and return rows with grade columns appended."""
    graded: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        question_preview = row['question'][:80]
        print(f"  [{index}/{len(rows)}] {question_preview}")

        eval_row = {
            'question': row['question'],
            'expectedResponse': row.get('expectedResponse', ''),
            'actualResponse': row.get('actualResponse', ''),
        }

        try:
            grades = grade_row(
                row=eval_row,
                judge_model=judge_model,
                judge_api_base=judge_api_base,
                judge_api_key=judge_api_key,
                temperature=temperature,
                max_tokens=max_tokens,
                allow_inline_citations=allow_inline_citations,
            )
        except Exception as exc:
            grades = {
                'generalQuality': 'Fail',
                'seemsRelevant': 'Fail',
                'seemsComplete': 'Skipped',
                'basedOnKnowledgeSources': 'Skipped',
                'reason': f'Judge error: {exc}',
            }

        graded.append({**row, **grades})
    return graded


def print_summary(all_results: dict[str, list[dict[str, str]]]) -> None:
    """Print a pass-rate summary table across all systems."""
    print('\n' + '=' * 70)
    print(f"{'System':<25} {'Pass':>6} {'Fail':>6} {'Total':>6} {'Rate':>8}")
    print('-' * 70)
    for sheet_name, rows in all_results.items():
        total = len(rows)
        passes = sum(1 for r in rows if r.get('generalQuality') == 'Pass')
        fails = total - passes
        rate = f'{passes / total:.0%}' if total else 'N/A'
        print(f'{sheet_name:<25} {passes:>6} {fails:>6} {total:>6} {rate:>8}')
    print('=' * 70)


def main() -> None:
    args = parse_args()

    if not args.input.exists():
        print(f'Error: input file not found: {args.input}', file=sys.stderr)
        sys.exit(1)

    judge_model = args.judge_model
    if not judge_model:
        try:
            health = fetch_yar_health(args.yar_api_url, args.yar_api_key)
            judge_model = str(health['configuration']['llm_model'])
        except Exception:
            judge_model = 'gpt-4.1-mini'
            print(f'Could not reach YAR /health; defaulting to judge model: {judge_model}')

    if not args.judge_api_key:
        print('Error: no judge API key. Set EVAL_LLM_BINDING_API_KEY in .env or pass --judge-api-key.', file=sys.stderr)
        sys.exit(1)

    output_path = args.output or timestamped_results_path('comparison_grades', '.xlsx')

    wb_in = openpyxl.load_workbook(args.input)
    target_sheets = (
        [s.strip() for s in args.sheets.split(',')]
        if args.sheets
        else [name for name in wb_in.sheetnames if name.lower() not in SKIP_SHEETS]
    )

    print(f'Input:  {args.input}')
    print(f'Output: {output_path}')
    print(f'Judge:  {normalize_litellm_model_name(judge_model, args.judge_api_base)}')
    print(f'Sheets: {", ".join(target_sheets)}')
    print()

    all_results: dict[str, list[dict[str, str]]] = {}
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default empty sheet

    for sheet_name in target_sheets:
        if sheet_name not in wb_in.sheetnames:
            print(f'Warning: sheet "{sheet_name}" not found, skipping.')
            continue

        ws = wb_in[sheet_name]
        rows = read_sheet_rows(ws)

        missing = REQUIRED_COLUMNS - set(rows[0].keys()) if rows else REQUIRED_COLUMNS
        if missing:
            print(f'Warning: sheet "{sheet_name}" missing columns {missing}, skipping.')
            continue

        print(f'--- Grading: {sheet_name} ({len(rows)} rows) ---')
        graded_rows = grade_sheet(
            sheet_name=sheet_name,
            rows=rows,
            judge_model=judge_model,
            judge_api_base=args.judge_api_base,
            judge_api_key=args.judge_api_key,
            temperature=args.temperature,
            max_tokens=args.max_tokens,
            allow_inline_citations=args.allow_inline_citations,
        )
        all_results[sheet_name] = graded_rows
        write_graded_sheet(wb_out, sheet_name, graded_rows)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    print(f'\nWrote: {output_path}')
    print_summary(all_results)


if __name__ == '__main__':
    main()
