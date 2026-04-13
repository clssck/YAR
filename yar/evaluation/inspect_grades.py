"""Dump grading results from a comparison grades XLSX to stdout.

Usage:
    python yar/evaluation/inspect_grades.py yar/evaluation/results/comparison_grades_*.xlsx
"""

import sys
from pathlib import Path

import openpyxl


def main() -> None:
    if len(sys.argv) < 2:
        print('Usage: python yar/evaluation/inspect_grades.py <grades.xlsx>')
        sys.exit(1)

    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path)

    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f'\n=== {name} ===')
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            d = dict(zip(headers, row, strict=False))
            q = (d.get('question') or '')[:70]
            gq = d.get('generalQuality', '?')
            reason = d.get('reason', '')
            print(f'  [{i}] {gq:>4}  {q}')
            print(f'        reason: {reason}')
        print()


if __name__ == '__main__':
    main()
